from openpyxl import load_workbook, Workbook

def LoadExcel(products,super):
    wb = Workbook()
    ws = wb.active

    ws["a1"] = "Marca"
    ws["b1"] = "Producto"
    ws["c1"] = "ValorActual"
    ws["d1"] = "ValorAntiguo"
    ws["e1"] = "Supermercado"
    ws["f1"] = "Tipo"
    ws["g1"] = "Promocion"


    #++++++++++++++++++++++++++++++++++++++++++++++++++++++
    #Esto es para insertar multiples filas por ejemplo si se requiere hacer ejecuciones de 10 hacer append a filas para luego recorrerlas y insertarlas
    #lista = [["Correo1","Api1",fecha],["Correo2","Api2",fecha]]

    for product in products:
        
        if product["actualPrice"] is None:
            product["actualPrice"] = "Agotado"

        if product["oldPrice"] is None:
            product["oldPrice"] = product["actualPrice"]
        if product["promotion"] is None:
            product["promotion"] = "No hay promocion"
            
        row = [product["brand"],product["name"],product["actualPrice"],product["oldPrice"],product["supermarket"],product["tipe"],product["promotion"]]
        ws.append(row)
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++
    wb.save(filename=f'informe{super}.xlsx')
    return True
